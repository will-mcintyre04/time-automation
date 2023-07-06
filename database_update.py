"""
Database Updater

Author: Will McIntyre
Date: 2023-07-05

Description:
This script provides a graphical user interface (GUI) for updating or uploading 
data to a time study database from a time tracking spreadsheet. The user specifies
the spreadsheet file and the desired database file to be updated or uploaded to.

Usage (See Documentation for further details):

Enter the path of the time tracking spreadsheet file into the first input.
Enter the path of the database file to be updated or leave blank to use the default (C:/Users/WMcIntyre/time_study.db)
Click the "Update Database" button to start the update process.
The script will validate the file format and the spreadsheet structure.
If the validation is successful, the data will be inserted or updated in the database.

Note: The default database path and other constants used in this script are specific to the author's directory
structure and can be modified according to your own requirements.
"""

import tkinter as tk
from tkinter import filedialog, messagebox
import sqlite3
import openpyxl

class DatabaseUpdater:
    DEFAULT_DB_PATH = "C:/Users/WMcIntyre/time_study.db"

    def __init__(self):
        self.window = None
        self.file_entry = None
        self.db_entry = None

    def upload_data_to_database(self, file_name, db_path):
        """ 
        Updates or uploads to a given time study database, giving each specific file a file_id, and updating/uploading
        file path, each action name, and the time associated with that action. The file inputted must match
        the expected time study format from "J:\\6.0 - Designer Folders\Will\Time Study Forms\Time Tracking Template v3.xlsm"
        """
        # Connect to the SQLite database
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        # Check if the file already exists
        cursor.execute("SELECT id FROM Files WHERE name = ?", (file_name,))
        existing_file_id = cursor.fetchone() #Returns first row as a tuple with id and name

        # If the file exists, prepare to update the data, if it does not, insert into the
        # "Files" table the new file and assign an id
        if existing_file_id:
            file_id = existing_file_id[0]
            # Delete the existing "Actions" data so we can update them
            cursor.execute("DELETE FROM Actions WHERE id = ?", (file_id, ))
        else:
            # Insert the file details into the "Files" table
            cursor.execute("INSERT INTO Files (name) VALUES (?)", (file_name,))
            file_id = cursor.lastrowid  # Get the auto-generated ID of the newly inserted row

        try:
            # Load the Excel sheet, without formulas
            workbook = openpyxl.load_workbook(file_name, data_only=True)
            sheet = workbook.active
        except openpyxl.utils.exceptions.InvalidFileException:
            # Print if error opening (invalid format or invalid entry name)
            messagebox.showwarning("Invalid File", "Invalid file format. Please ensure the file can be opened with Excel first.")
            return

        # Check the header values
        action_header = sheet.cell(row=3, column=2).value
        seconds_header = sheet.cell(row=3, column=4).value

        if action_header != "Action" or seconds_header != "Seconds":
            messagebox.showerror("Invalid Format", "The Excel sheet format is incorrect.")
            conn.close()
            return

        # Get the data range
        data_range = sheet.iter_rows(min_row=4, values_only=True)

        # Extract and insert the action data into the "Actions" table
        for row_index, row in enumerate(data_range, start=4):
            action_name = row[1]
            time_formula_cell = sheet.cell(row=row_index, column=4)
            time = time_formula_cell.value

            # Check for empty action column (column B), indicating the end of the table
            if action_name is None:
                self.window.destroy()
                break

            cursor.execute("INSERT INTO Actions (id, action_name, time) VALUES (?, ?, ?)", (file_id, action_name, time))

        # Commit the changes and close the connection
        conn.commit()
        conn.close()

        success_message = f"Database {db_path} has been updated with data from {file_name}"
        messagebox.showinfo("Success", success_message)

    def browse_file(self, specifier):
        """
        Allows user to browse for a file (spreadsheet or database) 
        and edits attributes dependant on the specifier (db or spread).
        """
        filetypes = [("Excel Files", "*.xlsm"), ("Database Files", "*.db")]
        file_path = filedialog.askopenfilename(filetypes=filetypes)
        if specifier == "spread":
            self.file_entry.delete(0, tk.END)
            self.file_entry.insert(tk.END, file_path)
        if specifier == "db":
            self.db_entry.delete(0, tk.END)
            self.db_entry.insert(tk.END, file_path)
    
    def destroy_window(self):
        """
        Exits the current window.
        """
        self.window.destroy()
    
    def main(self):
        """ 
        Populates a data updating GUI window for user input. 
        """
        self.window = tk.Tk()
        self.window.title("Excel to Database Upload")
        self.window.geometry("400x250+760+400")

        file_label= tk.Label(self.window, text= "Specify the spreadsheet to be added to the database:")
        file_label.pack()
        self.file_entry = tk.Entry(self.window)
        self.file_entry.pack()
        browse_button_spread = tk.Button(self.window, text="Browse", command=lambda: self.browse_file("spread"))
        browse_button_spread.pack()

        db_label = tk.Label(self.window, text="Specify the database to be updated:")
        db_label.pack()
        self.db_entry = tk.Entry(self.window, width=40)
        self.db_entry.insert(tk.END, self.DEFAULT_DB_PATH)
        self.db_entry.pack()
        browse_button_db = tk.Button(self.window, text="Browse", command=lambda: self.browse_file("db"))
        browse_button_db.pack()

        submit_button = tk.Button(self.window, text = "Update Database", 
                                command=lambda: self.upload_data_to_database(self.file_entry.get(), self.db_entry.get()),
                                bg = "#72B5E0", cursor = "hand2")

        submit_button.pack(pady=(20, 0))
        exit_button = tk.Button(self.window, text = "Exit", bg = "#FB5858", command=self.destroy_window, cursor="hand2")
        exit_button.pack(pady=(5, 0))

        self.window.mainloop()

if __name__ == "__main__":
    creator = DatabaseUpdater()
    creator.main()